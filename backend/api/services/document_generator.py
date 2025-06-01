# api/services/document_generator.py (новый код)
import os
import re
from docxtpl import DocxTemplate
import openpyxl
from django.conf import settings
from django.utils.dateformat import format as format_date

# Вспомогательная функция для получения вложенных значений из контекста
def get_value_from_context(context, key_string, default=''):
    keys = key_string.strip().split('.')
    value = context
    try:
        for key in keys:
            if isinstance(value, dict):
                value = value.get(key, default)
            else:
                value = getattr(value, key, default)
            if value is None: # Если на каком-то этапе None, возвращаем default
                 return default
        # Пробуем форматировать дату, если это объект даты/времени
        if hasattr(value, 'strftime'):
            try:
                # Используем Django's format для локализации, если возможно
                # или стандартный strftime
                return format_date(value, settings.DATE_FORMAT) # Или '%d.%m.%Y'
            except Exception:
                return str(value) # Возвращаем как строку, если форматирование не удалось
        return value if value is not None else default
    except (AttributeError, KeyError, TypeError):
        return default

def generate_document(order, document_type):
    TEMPLATE_MAP = {
        'contract': 'Договор.docx',
        'invoice': 'invoice_template.docx',
        'act': 'АКТ.xlsx',
        'transport_order': 'Заявка.docx',
        'additional_agreement': 'ЛАС Договор ТЭО.docx',
        'invoice_excel': 'ОТЧЕТ НУБИ.xlsx',
    }

    template_filename = TEMPLATE_MAP.get(document_type)
    if not template_filename:
        raise ValueError(f"Не найден шаблон для типа документа: {document_type}")

    template_path = os.path.join(
        settings.BASE_DIR,
        'templates/documents',
        template_filename
    )

    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Файл шаблона не найден: {template_path}")

    context = {
        'order': order,
        'client': order.client if order.client else None,
        'carrier': order.carrier if order.carrier else None,
        'contract_number': order.contract_number,
        'total_price': order.total_price,
        'unloading_date_formatted': order.unloading_date.strftime("%d.%m.%Y") if order.unloading_date else "Не указан",
        'contract_date_formatted': order.contract_date.strftime("%d.%m.%Y") if order.contract_date else "",
        'loading_date_formatted': order.loading_date.strftime("%d.%m.%Y %H:%M") if order.loading_date else "",
        'departure_date_formatted': order.departure_date.strftime("%d.%m.%Y %H:%M") if order.departure_date else "",
    }

    output_dir = os.path.join(settings.MEDIA_ROOT, 'documents', str(order.id))
    os.makedirs(output_dir, exist_ok=True)
    file_extension = os.path.splitext(template_filename)[1]
    output_filename = f"{document_type}_order_{order.id}{file_extension}"
    output_path = os.path.join(output_dir, output_filename)

    if file_extension == '.docx':
        try:
            doc = DocxTemplate(template_path)
            if not context['client']:
                 raise ValueError("Клиент не указан в заказе.")
            doc.render(context)
            doc.save(output_path)
        except Exception as e:
            print(f"Ошибка при генерации DOCX ({document_type}, Order ID: {order.id}): {e}")
            raise Exception(f"Ошибка генерации DOCX: {e}") from e

    elif file_extension == '.xlsx':
        try:
            workbook = openpyxl.load_workbook(template_path)
            placeholder_pattern = re.compile(r'\{\{\s*([\w\.]+)\s*\}\}')

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str):
                            original_value = cell.value
                            new_value_str = original_value
                            for match in placeholder_pattern.finditer(original_value):
                                key_string = match.group(1)
                                replacement_value = get_value_from_context(context, key_string)
                                new_value_str = new_value_str.replace(match.group(0), str(replacement_value))

                            if new_value_str != original_value:
                                if placeholder_pattern.fullmatch(original_value):
                                     final_value = get_value_from_context(context, placeholder_pattern.fullmatch(original_value).group(1))
                                     cell.value = final_value if final_value is not None else ''
                                else:
                                     cell.value = new_value_str

            workbook.save(output_path)
        except Exception as e:
             print(f"Ошибка при генерации XLSX ({document_type}, Order ID: {order.id}): {e}")
             raise Exception(f"Ошибка генерации XLSX: {e}") from e
    else:
        raise ValueError(f"Неподдерживаемый тип файла шаблона: {file_extension}")

    return output_path